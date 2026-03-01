"""Excel tracker generation — generates release-tracker.xlsx from the registry."""

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from apts_release.registry import load_registry
from apts_release.utils import format_size

# Styles
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
_ALT_ROW_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
_THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

_HEADERS = [
    "Release ID",
    "Date",
    "Product",
    "Release Name",
    "Release Version",
    "ESP32 Ver",
    "STM32 Ver",
    "WebUI Ver",
    "HMI Ver",
    "RPI Package",
    "RPI Size",
    "OTA Package",
    "OTA Size",
    "Notes",
    "Released By",
]

_COL_WIDTHS = [16, 12, 20, 30, 14, 14, 14, 14, 14, 42, 10, 42, 10, 50, 14]


def _format_date(timestamp: str) -> str:
    """Convert ISO timestamp to DD-MM-YYYY."""
    date_part = timestamp[:10]
    if len(date_part) == 10 and date_part[4] == "-":
        return f"{date_part[8:10]}-{date_part[5:7]}-{date_part[0:4]}"
    return date_part


def _release_to_row(release: dict) -> list[str]:
    """Convert a release entry dict to a flat row of display values."""
    comps = release.get("components", {})
    pkgs = release.get("packages", {})
    return [
        release.get("id", ""),
        _format_date(release.get("timestamp", "")),
        release.get("product", ""),
        release.get("release_name", ""),
        release.get("release_version", ""),
        comps.get("esp32_firmware", {}).get("version", ""),
        comps.get("stm32_firmware", {}).get("version", ""),
        comps.get("webpage", {}).get("version", ""),
        comps.get("hmi", {}).get("version") or "N/A",
        pkgs.get("rpi", {}).get("filename", "---"),
        format_size(pkgs.get("rpi", {}).get("size_bytes", 0)),
        pkgs.get("ota", {}).get("filename", "---"),
        format_size(pkgs.get("ota", {}).get("size_bytes", 0)),
        release.get("notes", ""),
        release.get("released_by", ""),
    ]


def _write_all_releases_sheet(ws, releases: list[dict]) -> None:
    """Populate the 'All Releases' sheet."""
    # Header row
    for col_idx, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    # Data rows (newest first)
    for row_idx, release in enumerate(reversed(releases), 2):
        row_data = _release_to_row(release)
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            if row_idx % 2 == 0:
                cell.fill = _ALT_ROW_FILL
            # Word-wrap for Notes column
            if col_idx == 14:
                cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Column widths
    for col_idx, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width


def _write_product_summary_sheet(ws, releases: list[dict]) -> None:
    """Populate the 'Product Summary' sheet."""
    # Gather per-product stats
    products: dict[str, list[dict]] = {}
    for r in releases:
        p = r.get("product", "UNKNOWN")
        products.setdefault(p, []).append(r)

    headers = [
        "Product",
        "Total Releases",
        "Latest Version",
        "Latest Date",
        "Latest ESP32",
        "Latest STM32",
        "First Release",
    ]

    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _THIN_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"

    row = 2
    for product_name, product_releases in sorted(products.items()):
        latest = product_releases[-1]
        earliest = product_releases[0]
        comps = latest.get("components", {})
        row_data = [
            product_name,
            len(product_releases),
            latest.get("release_version", ""),
            _format_date(latest.get("timestamp", "")),
            comps.get("esp32_firmware", {}).get("version", ""),
            comps.get("stm32_firmware", {}).get("version", ""),
            _format_date(earliest.get("timestamp", "")),
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row, column=col_idx, value=value)
            cell.border = _THIN_BORDER
            if row % 2 == 0:
                cell.fill = _ALT_ROW_FILL
        row += 1

    widths = [20, 14, 14, 14, 14, 14, 14]
    for col_idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = w


def generate_excel(registry_path: Path, output_path: Path) -> int:
    """Regenerate the Excel tracker from the registry. Returns row count."""
    registry = load_registry(registry_path)
    releases = registry.get("releases", [])

    wb = Workbook()

    # Sheet 1: All Releases
    ws_all = wb.active
    ws_all.title = "All Releases"
    _write_all_releases_sheet(ws_all, releases)

    # Sheet 2: Product Summary (always add it for consistency)
    ws_summary = wb.create_sheet("Product Summary")
    _write_product_summary_sheet(ws_summary, releases)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return len(releases)
